# This script is used to create an Excel file based on the cells found from the
# rule-based method.

from get_cell_coords import get_col_row, parse_coords, intersection_area  # Import necessary functions for cell coordinates and intersection calculation
from postprocess import find_loghi_words  # Import function for finding words in XML

import xml.etree.ElementTree as ET  # Import XML parsing library
import os  # Import OS module for file and directory operations
import cv2  # Import OpenCV for image processing
import numpy as np  # Import NumPy for numerical operations
import pandas as pd  # Import Pandas for data manipulation
from shapely.geometry import Point  # Import Point class for geometric operations
from shapely.geometry.polygon import Polygon  # Import Polygon class for geometric operations
from openpyxl import Workbook  # Import Workbook class for Excel operations
from openpyxl.drawing.image import Image as ExcelImage  # Import Image class for inserting images into Excel

def extract_col_row(xml_loghi, image_path):
    """
    Extracts cell polygons from an image using the XML file and the image path.

    Args:
    - xml_loghi (str): Path to the XML file containing information about the cells.
    - image_path (str): Path to the image.

    Returns:
    - List of shapely Polygon objects representing cells.
    """
    return get_col_row(xml_loghi, image_path)

def extract_text_data(xml_loghi):
    """
    Extracts text data along with coordinates from an XML file.

    Args:
    - xml_loghi (str): Path to the XML file.

    Returns:
    - Dictionary mapping word IDs to dictionaries containing 'coords' and 'text' keys.
    """
    return find_loghi_words(xml_loghi)

def assign_text_to_cells(column_coords, row_coords, loghi_words_dict):
    """
    Assigns text data to cells based on their coordinates.

    Args:
    - cells (list): List of shapely Polygon objects representing cells.
    - text_data (dict): Dictionary mapping word IDs to dictionaries containing text and coordinates.

    Returns:
    - Dictionary mapping words to cell IDs.
    """
    # Sort columns and rows by their bounding box coordinates
    sorted_columns = sorted(column_coords, key=lambda c: cv2.boundingRect(c)[0])
    sorted_rows = sorted(row_coords, key=lambda c: cv2.boundingRect(c)[1])
    sorted_loghi_words_dict = sorted(loghi_words_dict.values(), key=lambda x: min(coord[0] for coord in x['coords']))

    # Initialize a matrix for storing words assigned to each cell
    matrix = [[[] for _ in range(len(sorted_columns))] for _ in range(len(sorted_rows))]

    # Assign words to cells based on maximum overlap
    for word_data in sorted_loghi_words_dict:
        coords = word_data['coords']
        points = parse_coords(coords)
        word_box = cv2.boundingRect(np.array(points, np.int32))

        # Calculate overlap with columns and rows
        column_overlap = [intersection_area(word_box, cv2.boundingRect(col)) for col in sorted_columns]
        row_overlap = [intersection_area(word_box, cv2.boundingRect(row)) for row in sorted_rows]

        # Find column and row with maximum overlap
        max_column_index = np.argmax(column_overlap)
        max_row_index = np.argmax(row_overlap)

        matrix[max_row_index][max_column_index].append(word_data['text'])

    # Convert matrix to DataFrame
    df = pd.DataFrame(matrix, columns=range(len(matrix[0])), index=range(len(matrix)))

    # Define a function to remove rows with empty lists
    def remove_empty_rows(table):
        return table[table.applymap(lambda x: isinstance(x, list) and len(x) > 0).any(axis=1)]

    # Split the DataFrame into two separate tables if there are more than 10 columns
    if df.shape[1] > 10:
        table1 = df.iloc[:, :df.shape[1]//2]  # First half columns
        table2 = df.iloc[:, df.shape[1]//2:]  # Last half columns

        # Remove empty rows from each table
        table1 = remove_empty_rows(table1).reset_index(drop=True)
        table2 = remove_empty_rows(table2).reset_index(drop=True)

        return [table1, table2]
    else:
        df = remove_empty_rows(df).reset_index(drop=True)
        return df

def create_excel_sheet(matrix, image_path, output_file):
    """
    Creates an Excel sheet from the provided matrix and inserts the image next to the table.

    Args:
    - matrix (list): 2D list representing the cells and their assigned words.
    - image_path (str): Path to the image file.
    - output_file (str): Path to the output Excel file.
    """
    wb = Workbook()
    ws = wb.active

    # Write the data from the matrix to the Excel sheet
    for row_index, row in enumerate(matrix, start=1):
        for col_index, cell in enumerate(row, start=1):
            ws.cell(row=row_index, column=col_index, value=', '.join(cell))

    # Insert the image next to the table
    img = ExcelImage(image_path)

    img.width = 800
    img.height = 1000

    ws.add_image(img, 'Q1')  # Adjust the cell coordinates as needed

    # Save the Excel file
    wb.save(output_file)

def process_folder(main_folder):
    """
    Process each subdirectory in the main folder.

    Args:
    - main_folder (str): Path to the main folder containing subdirectories.
    """
    for subdir in os.listdir(main_folder):
        subdir_path = os.path.join(main_folder, subdir)
        if os.path.isdir(subdir_path):
            process_subdirectory(subdir_path)

def process_subdirectory(subdir_path):
    """
    Process a single subdirectory.

    Args:
    - subdir_path (str): Path to the subdirectory containing XML and JPG files.
    """
    page_folder = os.path.join(subdir_path, 'page')
    if not os.path.isdir(page_folder):
        return  # Skip if 'page' folder does not exist
    xml_files = []
    jpg_files = []
    for file in os.listdir(page_folder):
        if file.endswith('.xml'):
            xml_files.append(file)
    for file in os.listdir(subdir_path):
        if file.endswith('.jpg'):
            jpg_files.append(file)

    for xml_file in xml_files:
        jpg_file = xml_file.replace('.xml', '.jpg')
        if jpg_file in jpg_files:
            xml_path = os.path.join(page_folder, xml_file)
            jpg_path = os.path.join(subdir_path, jpg_file)
            process_files(xml_path, jpg_path)

def process_files(xml_path, jpg_path):
    """
    Process a pair of XML and JPG files.

    Args:
    - xml_path (str): Path to the XML file.
    - jpg_path (str): Path to the JPG file.
    """
    col, row = extract_col_row(xml_path, jpg_path)
    text_data = extract_text_data(xml_path)
    word_cell_assignments = assign_text_to_cells(col, row, text_data)
    output_file = os.path.splitext(os.path.basename(xml_path))[0] + '_output.xlsx'
    create_excel_sheet(word_cell_assignments, jpg_path, '/home/roderickmajoor/Desktop/Master/Thesis/Excel/' + output_file)

def create_excel_workbook(matrix, image_path, wb):
    """
    Creates an Excel sheet with the provided matrix and inserts the resized image at the specified position.

    Args:
    - matrix (list): 2D list representing the cells and their assigned words.
    - image_path (str): Path to the image file.
    - wb (openpyxl.Workbook): Excel workbook object to which the sheet will be added.
    """
    ws = wb.create_sheet(title=os.path.basename(image_path))  # Create a sheet with image name
    # Write the data from the matrix to the Excel sheet

    if isinstance(matrix, list):
        table1, table2 = matrix

        for row_index, row in table1.iterrows():
            for col_index, cell in enumerate(row):
                if cell:
                    ws.cell(row=row_index + 1, column=col_index + 1, value=', '.join(cell))

        for row_index, row in table2.iterrows():
            for col_index, cell in enumerate(row):
                if cell:
                    ws.cell(row=row_index + 1, column=col_index + 2 + table1.shape[1], value=', '.join(cell))
    else:
        df = matrix
        for row_index, row in df.iterrows():
            for col_index, cell in enumerate(row):
                if cell:
                    ws.cell(row=row_index + 1, column=col_index + 1, value=', '.join(cell))

    # Insert the image next to the table
    img = ExcelImage(image_path)

    img.width = 800
    img.height = 1000

    ws.add_image(img, 'S1')  # Adjust the cell coordinates as needed

def create_excel_file(main_folder, output_file):
    """
    Creates an Excel file with different sheets for each image/XML pair.

    Args:
    - main_folder (str): Path to the main folder containing subdirectories.
    - output_file (str): Path to the output Excel file.
    """
    wb = Workbook()

    for subdir in os.listdir(main_folder):
        print("Processing subdir: " + subdir)
        subdir_path = os.path.join(main_folder, subdir)
        if os.path.isdir(subdir_path):
            page_folder = os.path.join(subdir_path, 'page')
            if not os.path.isdir(page_folder):
                continue  # Skip if 'page' folder does not exist

            xml_files = [f for f in os.listdir(page_folder) if f.endswith('.xml')]
            jpg_files = [f for f in os.listdir(subdir_path) if f.endswith('.jpg')]

            for xml_file in xml_files:
                jpg_file = xml_file.replace('.xml', '.jpg')
                if jpg_file in jpg_files:
                    xml_path = os.path.join(page_folder, xml_file)
                    jpg_path = os.path.join(subdir_path, jpg_file)
                    column_coords, row_coords = extract_col_row(xml_path, jpg_path)
                    loghi_words_dict = find_loghi_words(xml_path)
                    matrix = assign_text_to_cells(column_coords, row_coords, loghi_words_dict)
                    create_excel_workbook(matrix, jpg_path, wb)

    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)

    # Save the Excel file
    wb.save(output_file)

# Example usage:
# Main folder is expected to be a folder containing subfolders.
# These subfolders are expected to have 1) .jpg files and 2) a folder called 'page' containing the corresponding loghi pageXML files.
main_folder = "/home/roderickmajoor/Desktop/Master/Thesis/loghi/data/"
output_file = "/home/roderickmajoor/Desktop/Master/Thesis/Excel/output1.xlsx"
create_excel_file(main_folder, output_file)

